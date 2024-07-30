from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Set up the Selenium webdriver (make sure you have the appropriate browser driver installed)
driver = (
    webdriver.Chrome()
)  # For Chrome, you can use webdriver.Firefox() for Firefox, etc.

# Send a request to the main page
main_url = "https://docs.aws.amazon.com/appsync/latest/APIReference/Welcome.html"
driver.get(main_url)

# Find the list items using the specified XPath
# where is are the urls on the side bar
list_items_xpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/nav[2]/div/div[2]/div/ul/li[2]/div/div[2]/ul"
list_items = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, list_items_xpath))
)
# Create a new Excel workbook
workbook = openpyxl.Workbook()
# Iterate over each list item
for item in list_items.find_elements(By.TAG_NAME, "li"):
    # Extract the URL from the list item
    item_url = item.find_element(By.TAG_NAME, "a").get_attribute("href")
    # Send a request to the item page
    driver.get(item_url)
    # Find the table using the specified XPath
    table_xpath = "/html/body/div[2]/div/div/div[3]/div/div/main/div/div/div[1]/div[1]/div/div/div[4]/div[1]/div[5]/div/table"
    table = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, table_xpath))
    )
    # Create a new sheet for the item
    sheet_name = item.text.strip()
    sheet = workbook.create_sheet(title=sheet_name)
    # Iterate over each row in the table and write the data to the sheet
    for row_index, row in enumerate(table.find_elements(By.TAG_NAME, "tr")):
        for col_index, cell in enumerate(
            row.find_elements(By.XPATH, "./*[self::th or self::td]")
        ):
            sheet.cell(row=row_index + 1, column=col_index + 1, value=cell.text.strip())
# Remove the default sheet created by openpyxl
# workbook.remove(workbook["Sheet"])
# Save the Excel workbook
# workbook.save("output.xlsx")
# Close the webdriver
driver.quit()
